import os
from typing import List, Dict, Any


def extract_text_from_pdf(file_path: str) -> List[Dict[str, Any]]:
    from pypdf import PdfReader
    reader = PdfReader(file_path)
    blocks = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            blocks.append({"page_num": i + 1, "text": text.strip()})
    return blocks


def extract_text_from_docx(file_path: str) -> List[Dict[str, Any]]:
    from docx import Document
    doc = Document(file_path)
    blocks = []
    current_section = "Document"
    for para in doc.paragraphs:
        if not para.text.strip():
            continue
        if para.style.name.startswith("Heading"):
            current_section = para.text.strip()
        blocks.append({"section": current_section, "text": para.text.strip()})
    return blocks


def extract_text_from_xlsx(file_path: str) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    blocks = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text = []
        for row in ws.iter_rows():
            cell_values = [str(cell.value) for cell in row if cell.value is not None]
            if cell_values:
                rows_text.append(" | ".join(cell_values))
        if rows_text:
            blocks.append({"sheet_name": sheet_name, "text": "\n".join(rows_text)})
    return blocks


def chunk_text(text_blocks: List[Dict[str, Any]], chunk_size: int = 800) -> List[Dict[str, Any]]:
    chunks = []
    for block in text_blocks:
        text = block.get("text", "")
        metadata = {k: v for k, v in block.items() if k != "text"}
        # Split into chunks
        words = text.split()
        current_chunk = []
        current_len = 0
        for word in words:
            current_chunk.append(word)
            current_len += len(word) + 1
            if current_len >= chunk_size:
                chunks.append({"text": " ".join(current_chunk), "metadata": metadata})
                current_chunk = []
                current_len = 0
        if current_chunk:
            chunks.append({"text": " ".join(current_chunk), "metadata": metadata})
    return chunks


def process_document(file_path: str, file_type: str) -> List[Dict[str, Any]]:
    if file_type == "pdf":
        blocks = extract_text_from_pdf(file_path)
    elif file_type == "docx":
        blocks = extract_text_from_docx(file_path)
    elif file_type == "xlsx":
        blocks = extract_text_from_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file type: {file_type}")
    return chunk_text(blocks)
